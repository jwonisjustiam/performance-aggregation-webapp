"""Flexible, cross-platform Excel input reader."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook

from services.file_security import readable_workbook

XLS_ERROR = "현재 버전에서는 .xlsx 파일만 지원합니다. .xls 파일은 xlsx로 변환 후 다시 업로드해주세요."

ALIASES = {
    "주문번호": ("주문번호", "상품주문번호", "order id"),
    "결제일시": ("결제일시", "결제일", "결제 일시"),
    "상품명": ("상품명", "상품 이름"),
    "수량": ("수량", "상품수량"),
    "옵션관리코드": ("옵션관리코드", "옵션 관리 코드"),
    "판매자 상품코드": ("판매자 상품코드", "판매자상품코드"),
    "상품가격": ("상품가격", "상품 가격"),
    "옵션가격": ("옵션가격", "옵션 가격"),
    "최종 상품별 총 주문금액": ("최종 상품별 총 주문금액", "최종상품별총주문금액"),
    "주문 유입경로": ("주문 유입경로", "주문유입경로", "쇼핑라이브 여부"),
}


@dataclass
class WorkbookData:
    data: pd.DataFrame
    sheet_names: list[str]
    selected_sheet: str
    header_row: int
    encrypted: bool


def normalize_label(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "")).lower()


def canonicalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Rename recognized source columns to canonical business names."""
    normalized = {normalize_label(column): column for column in df.columns}
    rename: dict[object, str] = {}
    for canonical, aliases in ALIASES.items():
        for alias in aliases:
            found = normalized.get(normalize_label(alias))
            if found is not None:
                rename[found] = canonical
                break
    return df.rename(columns=rename)


def detect_header_row(path: Path, sheet_name: str, scan_rows: int = 30) -> int:
    """Find the most likely header row using recognized column aliases."""
    preview = pd.read_excel(path, sheet_name=sheet_name, header=None, nrows=scan_rows, engine="openpyxl")
    known = {normalize_label(alias) for aliases in ALIASES.values() for alias in aliases}
    scores = []
    for index, row in preview.iterrows():
        score = sum(normalize_label(value) in known for value in row if pd.notna(value))
        scores.append((score, int(index)))
    best_score, best_index = max(scores, default=(0, 0))
    if best_score < 2:
        return 0
    return best_index


def read_xlsx(path: Path, sheet_name: str | None = None) -> WorkbookData:
    """Read an xlsx workbook, including supported encrypted files."""
    if path.suffix.lower() != ".xlsx":
        raise ValueError(XLS_ERROR)
    with readable_workbook(path) as (readable, encrypted):
        workbook = load_workbook(readable, read_only=True, data_only=False)
        try:
            sheets = list(workbook.sheetnames)
        finally:
            workbook.close()
        selected = sheet_name if sheet_name in sheets else sheets[0]
        header_row = detect_header_row(readable, selected)
        frame = pd.read_excel(readable, sheet_name=selected, header=header_row, engine="openpyxl")
    frame = frame.dropna(how="all").reset_index(drop=True)
    return WorkbookData(canonicalize_columns(frame), sheets, selected, header_row + 1, encrypted)


def first_present(columns: Iterable[object], candidates: Iterable[str]) -> str | None:
    lookup = {normalize_label(column): str(column) for column in columns}
    for candidate in candidates:
        if normalize_label(candidate) in lookup:
            return lookup[normalize_label(candidate)]
    return None

