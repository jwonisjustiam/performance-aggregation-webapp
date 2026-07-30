"""Flexible multi-market Excel input reader."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook

from services.file_security import readable_workbook
from services.sku_resolver import extract_sku_code, normalize_sku

SUPPORTED_EXTENSIONS = {".xlsx", ".xls"}
XLS_ERROR = "지원 형식은 .xlsx와 .xls입니다."

ALIASES = {
    "주문번호": ("주문번호", "order id"),
    "상품주문번호": ("상품주문번호", "상품 주문번호", "주문상세번호"),
    "결제일시": (
        "결제일시",
        "결제일",
        "결제 일시",
        "예약결제완료일시",
        "주문일시",
        "주문일",
        "주문일자",
    ),
    "상품명": ("상품명", "상품 이름"),
    "수량": ("수량", "상품수량"),
    "옵션 정보": ("옵션 정보", "옵션"),
    "옵션관리코드": (
        "옵션관리코드",
        "옵션 관리 코드",
        "옵션코드",
        "상품옵션코드",
        "상품 옵션코드",
        "옵션코드번호",
        "옵션 코드 번호",
        "판매자상세관리코드",
        "판매자 상세관리코드",
    ),
    "판매자 상품코드": (
        "판매자 상품코드",
        "판매자상품코드",
        "판매자관리코드",
        "판매자 관리코드",
        "판매자상품번호",
    ),
    "상품가격": ("상품가격", "상품 가격", "상품금액", "판매단가"),
    "옵션가격": ("옵션가격", "옵션 가격", "옵션금액", "옵션가"),
    "최종 상품별 총 주문금액": (
        "최종 상품별 총 주문금액",
        "최종상품별총주문금액",
        "주문금액",
        "판매금액",
        "구매금액",
    ),
    "주문 유입경로": ("주문 유입경로", "주문유입경로", "쇼핑라이브 여부"),
}

MARKET_MARKERS = {
    "11번가": {"예약결제완료일시", "주문상세번호", "서비스이용료정책"},
    "카카오": {"결제번호", "채널상품번호", "톡딜여부", "biz판매여부"},
    "지마켓/옥션": {"판매아이디", "판매자관리코드", "판매자 관리코드", "ssg 원주문번호"},
    "네이버": {"상품주문번호", "주문 유입경로", "옵션관리코드"},
}


@dataclass
class WorkbookData:
    data: pd.DataFrame
    sheet_names: list[str]
    selected_sheet: str
    header_row: int | str
    encrypted: bool


def normalize_label(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "")).lower()


def _first_nonempty(row: pd.Series, candidates: Iterable[str]) -> str:
    for candidate in candidates:
        if candidate not in row.index:
            continue
        value = row.get(candidate)
        if value is not None and not pd.isna(value) and str(value).strip():
            return str(value).strip()
    return ""


def detect_market(columns: Iterable[object], source_name: str = "") -> str:
    """Detect the source mall from unique headers, then fall back to the file name."""
    normalized = {normalize_label(column) for column in columns}
    scores = {
        market: len(normalized.intersection({normalize_label(marker) for marker in markers}))
        for market, markers in MARKET_MARKERS.items()
    }
    market, score = max(scores.items(), key=lambda item: item[1])
    if score:
        return market
    lowered_name = source_name.lower()
    if "11번가" in lowered_name or "11st" in lowered_name:
        return "11번가"
    if "카카오" in lowered_name or "kakao" in lowered_name:
        return "카카오"
    if "옥션" in lowered_name or "auction" in lowered_name:
        return "옥션"
    if "지마켓" in lowered_name or "gmarket" in lowered_name:
        return "지마켓"
    if "네이버" in lowered_name or "naver" in lowered_name:
        return "네이버"
    return "미확인"


def _platform_for_row(row: pd.Series, detected_market: str) -> str:
    seller_id = _first_nonempty(row, ("판매아이디",))
    if "옥션" in seller_id:
        return "옥션"
    if "지마켓" in seller_id:
        return "지마켓"
    return detected_market


def _derive_sku(row: pd.Series) -> str:
    existing = _first_nonempty(row, ("옵션관리코드",))
    existing_sku = extract_sku_code(existing)
    if existing_sku:
        return existing_sku

    seller_code = _first_nonempty(row, ("판매자 상품코드",))
    product_name = _first_nonempty(row, ("상품명",))
    base_sku = extract_sku_code(seller_code, product_name)
    option_text = _first_nonempty(row, ("옵션 정보", "옵션"))
    suffix_match = re.search(r"[\{\[]([A-Z0-9-]+)[\}\]]", option_text, flags=re.IGNORECASE)
    if base_sku and suffix_match:
        base = normalize_sku(base_sku)
        suffix = suffix_match.group(1).upper()
        return base if base.endswith(suffix) else f"{base}{suffix}"
    if base_sku:
        return base_sku
    return existing or seller_code


def canonicalize_columns(df: pd.DataFrame, source_name: str = "") -> pd.DataFrame:
    """Rename recognized source columns and derive fields shared by all processors."""
    original_columns = list(df.columns)
    detected_market = detect_market(original_columns, source_name)
    normalized = {normalize_label(column): column for column in original_columns}
    rename: dict[object, str] = {}
    claimed: set[str] = set()
    for canonical, aliases in ALIASES.items():
        for alias in aliases:
            found = normalized.get(normalize_label(alias))
            if found is not None and canonical not in claimed:
                rename[found] = canonical
                claimed.add(canonical)
                break

    result = df.rename(columns=rename).copy()
    if "주문번호" not in result and "상품주문번호" in result:
        result["주문번호"] = result["상품주문번호"]
    if "주문번호" not in result and "결제번호" in result:
        result["주문번호"] = result["결제번호"]

    result["원본몰"] = result.apply(lambda row: _platform_for_row(row, detected_market), axis=1)
    derived_skus = result.apply(_derive_sku, axis=1)
    if "옵션관리코드" in result:
        result["옵션관리코드"] = result["옵션관리코드"].astype(object)
        blank = result["옵션관리코드"].isna() | result["옵션관리코드"].astype(str).str.strip().eq("")
        result.loc[blank, "옵션관리코드"] = derived_skus.loc[blank]
    else:
        result["옵션관리코드"] = derived_skus

    if "주문 유입경로" in result:
        result["쇼핑라이브 판정근거"] = "원본 구분 열"
    else:
        result["쇼핑라이브 판정근거"] = "원본 구분 열 없음—업로드 행 전체를 후보로 사용"
    return result


def _engine_for(path: Path) -> str:
    return "xlrd" if path.suffix.lower() == ".xls" else "openpyxl"


def _sheet_names(path: Path) -> list[str]:
    if path.suffix.lower() == ".xls":
        return list(pd.ExcelFile(path, engine="xlrd").sheet_names)
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _detect_header_row_with_score(path: Path, sheet_name: str, scan_rows: int = 40) -> tuple[int, int]:
    """Find the most likely header row using aliases from every supported mall."""
    preview = pd.read_excel(path, sheet_name=sheet_name, header=None, nrows=scan_rows, engine=_engine_for(path))
    known = {normalize_label(alias) for aliases in ALIASES.values() for alias in aliases}
    known.update(normalize_label(marker) for markers in MARKET_MARKERS.values() for marker in markers)
    scores = []
    for index, row in preview.iterrows():
        score = sum(normalize_label(value) in known for value in row if pd.notna(value))
        scores.append((score, int(index)))
    best_score, best_index = max(scores, default=(0, 0))
    if best_score < 2:
        return 0, best_score
    return best_index, best_score


def detect_header_row(path: Path, sheet_name: str, scan_rows: int = 40) -> int:
    header_row, _ = _detect_header_row_with_score(path, sheet_name, scan_rows)
    return header_row


def _looks_like_order_sheet(frame: pd.DataFrame) -> bool:
    recognized = {
        "주문번호",
        "상품주문번호",
        "결제일시",
        "상품명",
        "수량",
        "옵션관리코드",
        "판매자 상품코드",
        "최종 상품별 총 주문금액",
    }
    present = recognized.intersection(set(map(str, frame.columns)))
    return len(present) >= 3 and not frame.dropna(how="all").empty


def _deduplicate_overlapping_sheets(frame: pd.DataFrame) -> pd.DataFrame:
    """Remove copies of the same order line kept in original and edited sheets."""
    if frame.empty or "주문번호" not in frame:
        return frame

    def identity(row: pd.Series) -> tuple[str, ...]:
        product_order = _first_nonempty(row, ("상품주문번호",))
        order_identity = product_order or _first_nonempty(row, ("주문번호",))
        if not order_identity:
            return ("__row__", str(row.name))
        return (
            _first_nonempty(row, ("원본몰",)),
            order_identity,
            _first_nonempty(row, ("상품명",)),
            _first_nonempty(row, ("옵션관리코드", "옵션 정보")),
            _first_nonempty(row, ("결제일시",)),
        )

    identities = frame.apply(identity, axis=1)
    return frame.loc[~identities.duplicated(keep="first")].reset_index(drop=True)


def read_workbook(path: Path, sheet_name: str | None = None) -> WorkbookData:
    """Read xlsx/xls workbooks, including supported password-protected files."""
    if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError(XLS_ERROR)
    with readable_workbook(path) as (readable, encrypted):
        sheets = _sheet_names(readable)
        target_sheets = [sheet_name] if sheet_name in sheets else sheets
        frames: list[tuple[int, int, int, pd.DataFrame]] = []
        selected_sheets: list[str] = []
        header_rows: list[str] = []
        for current_sheet in target_sheets:
            header_row, header_score = _detect_header_row_with_score(readable, current_sheet)
            frame = pd.read_excel(readable, sheet_name=current_sheet, header=header_row, engine=_engine_for(readable))
            frame = canonicalize_columns(frame.dropna(how="all").reset_index(drop=True), path.name)
            if sheet_name in sheets or header_score >= 2 and _looks_like_order_sheet(frame):
                frame["원본시트"] = current_sheet
                frames.append((header_score, len(frame.columns), len(frame), frame))
                selected_sheets.append(current_sheet)
                header_rows.append(f"{current_sheet}:{header_row + 1}")

        if frames:
            ranked_frames = [item[3] for item in sorted(frames, key=lambda item: item[:3], reverse=True)]
            combined = pd.concat(ranked_frames, ignore_index=True, sort=False)
            combined = _deduplicate_overlapping_sheets(combined)
            selected = ", ".join(selected_sheets)
            header_row_text: int | str = (
                int(header_rows[0].rsplit(":", 1)[1])
                if len(header_rows) == 1
                else ", ".join(header_rows)
            )
        else:
            selected = sheets[0]
            header_row = detect_header_row(readable, selected)
            combined = pd.read_excel(readable, sheet_name=selected, header=header_row, engine=_engine_for(readable))
            combined = canonicalize_columns(combined.dropna(how="all").reset_index(drop=True), path.name)
            header_row_text = header_row + 1
    return WorkbookData(combined, sheets, selected, header_row_text, encrypted)


def read_xlsx(path: Path, sheet_name: str | None = None) -> WorkbookData:
    """Backward-compatible alias for callers that still import read_xlsx."""
    return read_workbook(path, sheet_name)


def first_present(columns: Iterable[object], candidates: Iterable[str]) -> str | None:
    lookup = {normalize_label(column): str(column) for column in columns}
    for candidate in candidates:
        if normalize_label(candidate) in lookup:
            return lookup[normalize_label(candidate)]
    return None
