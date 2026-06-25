"""Generate styled, validated result workbooks."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
import tempfile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from services.validator import validate_saved_workbook


def _style(path: Path, sheet_formats: dict[str, dict[str, str]]) -> None:
    workbook = load_workbook(path)
    try:
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            sheet.sheet_view.showGridLines = False
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
            for column_cells in sheet.columns:
                width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 45)
                sheet.column_dimensions[column_cells[0].column_letter].width = max(width, 10)
            formats = sheet_formats.get(sheet.title, {})
            header_map = {cell.value: cell.column for cell in sheet[1]}
            for header, number_format in formats.items():
                if header in header_map:
                    for row in range(2, sheet.max_row + 1):
                        sheet.cell(row, header_map[header]).number_format = number_format
        workbook.save(path)
    finally:
        workbook.close()


def create_result_workbook(job_type: str, result: dict[str, pd.DataFrame]) -> tuple[bytes, dict[str, object]]:
    """Create the required workbook, reopen it, and return bytes plus validation."""
    required = ["회차별 합계"] if job_type == "weekly" else ["통합 실적표", "회차별 합계", "중복 주문 검증"]
    with tempfile.TemporaryDirectory() as temporary:
        path = Path(temporary) / "result.xlsx"
        with pd.ExcelWriter(path, engine="xlsxwriter", datetime_format="yyyy-mm-dd") as writer:
            if job_type == "weekly":
                result["final"].to_excel(writer, sheet_name="회차별 합계", index=False)
            else:
                result["final"].to_excel(writer, sheet_name="통합 실적표", index=False)
                result["summary"].to_excel(writer, sheet_name="회차별 합계", index=False)
                result["duplicates"].to_excel(writer, sheet_name="중복 주문 검증", index=False)
        _style(
            path,
            {
                "회차별 합계": {"금액(백만)": "0.###", "총 금액": "#,##0", "수량": "0", "전환율": "0"},
                "통합 실적표": {"실적(대)": "0", "Duration (분)": "0", "View(만)": "0.###"},
                "중복 주문 검증": {"주문 금액": "#,##0"},
            },
        )
        validation = validate_saved_workbook(path, required)
        if not validation["valid"]:
            raise ValueError(f"결과 파일 검증 실패: {validation}")
        return path.read_bytes(), validation

