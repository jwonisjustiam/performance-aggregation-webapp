"""Input and output validation."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook

FORMULA_ERRORS = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}


def missing_columns(df: pd.DataFrame, required: Iterable[str]) -> list[str]:
    return [column for column in required if column not in df.columns]


def input_diagnostics(df: pd.DataFrame, required: Iterable[str]) -> dict[str, int | bool | list[str]]:
    """Return compact input-quality diagnostics."""
    missing = missing_columns(df, required)
    dates = pd.to_datetime(df.get("결제일시"), errors="coerce") if "결제일시" in df else pd.Series(dtype="datetime64[ns]")
    quantities = pd.to_numeric(df.get("수량"), errors="coerce") if "수량" in df else pd.Series(dtype=float)
    amount_columns = [column for column in ("최종 상품별 총 주문금액", "상품가격", "옵션가격") if column in df]
    amount_errors = 0
    for column in amount_columns:
        amount_errors += int(pd.to_numeric(df[column].astype(str).str.replace(",", "", regex=False), errors="coerce").isna().sum())
    required_empty = sum(int(df[column].isna().sum()) for column in required if column in df)
    duplicate_orders = int(df["주문번호"].duplicated().sum()) if "주문번호" in df else 0
    shifted_live = any(
        str(value).strip() == "0" and index + 1 < len(row) and str(row.iloc[index + 1]).strip() == "쇼핑라이브"
        for _, row in df.iterrows()
        for index, value in enumerate(row)
    )
    return {
        "missing_columns": missing,
        "date_errors": int(dates.isna().sum()) if len(dates) else 0,
        "quantity_errors": int(quantities.isna().sum()) if len(quantities) else 0,
        "amount_errors": amount_errors,
        "duplicate_orders": duplicate_orders,
        "required_empty": required_empty,
        "shifted_columns_detected": shifted_live,
        "live_source_decidable": "주문 유입경로" in df.columns or shifted_live,
    }


def validate_saved_workbook(path: Path, required_sheets: Iterable[str]) -> dict[str, object]:
    """Reopen an output workbook and scan for sheet and formula errors."""
    workbook = load_workbook(path, data_only=False)
    try:
        names = workbook.sheetnames
        missing = [name for name in required_sheets if name not in names]
        errors: list[str] = []
        empty: list[str] = []
        for sheet in workbook.worksheets:
            if sheet.max_row < 2:
                empty.append(sheet.title)
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value in FORMULA_ERRORS:
                        errors.append(f"{sheet.title}!{cell.coordinate}:{cell.value}")
        return {"valid": not missing and not errors and not empty, "missing_sheets": missing, "formula_errors": errors, "empty_sheets": empty}
    finally:
        workbook.close()

