from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from processors.weekly_processor import process_weekly
from services.excel_writer import create_result_workbook


ROOT = Path(__file__).parents[1]


def test_streamlit_entrypoint_and_docs() -> None:
    app_source = (ROOT / "app.py").read_text(encoding="utf-8")
    requirements = (ROOT / "requirements.txt").read_text(encoding="utf-8")
    readme = (ROOT / "README.md").read_text(encoding="utf-8")

    assert "import streamlit as st" in app_source
    assert "사용 안내" in app_source
    assert "JOB_TYPE_OPTIONS" in app_source
    assert "삼성 취합" in app_source
    assert "위클리 취합" in app_source
    assert "워치9 사전판매 판매 실적 취합" in app_source
    assert "selected_job[\"title\"]" in app_source
    assert "raw_files_{job_type}" in app_source
    assert "방송 실적표" not in app_source
    assert "네이버 API 자동 수집" not in app_source
    assert "fetch_raw_data" not in app_source
    assert "streamlit==" in requirements
    assert "fastapi" not in requirements.lower()
    assert "requests==" not in requirements.lower()
    assert "bcrypt==" not in requirements.lower()
    assert "Streamlit Cloud" in readme


def test_weekly_sample_can_create_downloadable_workbook() -> None:
    raw = pd.DataFrame(
        [["A", "2026-07-07 14:00", "외장하드", "옵션", 1_000_000, 0, "쇼핑라이브"]],
        columns=["주문번호", "결제일시", "상품명", "옵션 정보", "상품가격", "옵션가격", "주문 유입경로"],
    )
    result = process_weekly(raw, "외장하드.xlsx", "external")
    content, validation = create_result_workbook("weekly", result)

    assert validation["valid"] is True

    path = ROOT / "outputs" / "_test_streamlit_weekly.xlsx"
    path.write_bytes(content)
    workbook = load_workbook(path, read_only=True)
    try:
        assert workbook.sheetnames == ["회차별 합계"]
    finally:
        workbook.close()
        path.unlink(missing_ok=True)


def test_detail_workbook_sheet_is_created() -> None:
    raw = pd.DataFrame(
        [["A", "2026-07-07 14:00", "갤럭시 워치9", "", "SELLER-A", 1_000_000, 0, "검색"]],
        columns=["주문번호", "결제일시", "상품명", "옵션관리코드", "판매자 상품코드", "상품가격", "옵션가격", "주문 유입경로"],
    )
    from processors.weekly_processor import process_detail

    result = process_detail(raw, "주문.xlsx")
    content, validation = create_result_workbook("detail", result)

    assert validation["valid"] is True

    path = ROOT / "outputs" / "_test_streamlit_detail.xlsx"
    path.write_bytes(content)
    workbook = load_workbook(path, read_only=True)
    try:
        assert workbook.sheetnames == ["웨어러블", "모바일 ACC", "전체 미리보기"]
    finally:
        workbook.close()
        path.unlink(missing_ok=True)
